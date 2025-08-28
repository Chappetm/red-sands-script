#!/usr/bin/env python3
import argparse
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Índices de columnas en el CSV
ROOM_IDX = 8      # Columna I
NAME_IDX = 10     # Columna K
SURNAME_IDX = 11  # Columna L
MEAL_IDX = 26     # Columna AA
STATUS_IDX = 4    # Columna E

STATUS_EXCLUDE = {"cancelled", "checked out"}  # case-insensitive

def today_str():
    return datetime.now().strftime("%d-%m-%Y")

def build_meal_list(df: pd.DataFrame) -> pd.DataFrame:
    surname_series = df.iloc[:, SURNAME_IDX].astype(str)
    status_series = df.iloc[:, STATUS_IDX].astype(str)

    # Filtrar STAFF, COSA y UNKNOWN (case-insensitive)
    mask_staff   = surname_series.str.upper().str.contains(r"\(STAFF\)", na=False)
    mask_cosa    = surname_series.str.upper().str.contains("COSA", na=False)
    mask_unknown = surname_series.str.upper().str.contains("UNKNOWN", na=False)

    # Filtrar status
    mask_status = status_series.str.strip().str.casefold().isin(STATUS_EXCLUDE)

    # Filtrar todo
    filtered = df[~(mask_staff | mask_cosa | mask_unknown | mask_status)].copy()

    # Subset columnas
    out = filtered.iloc[:, [ROOM_IDX, NAME_IDX, SURNAME_IDX, MEAL_IDX]].copy()
    out.columns = ["Room", "Name", "Surname", "Meal option"]

    # Limpieza de espacios
    for c in ["Room", "Name", "Surname"]:
        out[c] = out[c].astype(str).str.strip()

    # Interpretar meal option desde las notas
    def detect_package(note: str) -> str:
        note_low = str(note).lower()
        if "meal package" in note_low:
            return "MEAL PACKAGE"
        elif "bill to room" in note_low:
            return "BILL TO ROOM"
        elif "room only" in note_low:
            return "ROOM ONLY"
        elif "breakfast only" in note_low or "breakfast" in note_low:
            return "BREAKFAST ONLY"
        else:
            return "NO INFO ⚠️"

    out["Meal option"] = out["Meal option"].apply(detect_package)

    out = out.sort_values(by="Room", key=lambda col: col.astype(str).str.lower()).reset_index(drop=True)
    return out

def write_excel(meals_df: pd.DataFrame, out_path: Path, date_text: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Meals"

    # Encabezados de Excel
    ws.cell(row=1, column=1).value = "MEALS LIST"
    ws.cell(row=2, column=1).value = date_text
    headers = ["Room", "Name", "Surname", "Meal option"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=3, column=j).value = h

    # Datos
    for i, r in enumerate(meals_df.itertuples(index=False), start=4):
        ws.cell(row=i, column=1).value = r[0]
        ws.cell(row=i, column=2).value = r[1]
        ws.cell(row=i, column=3).value = r[2]
        ws.cell(row=i, column=4).value = r[3]

    # Anchos de columnas
    widths = [12, 22, 22, 24]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # --- FORMATO ---
    # Merge de título y fecha
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')

    # Fuente y alineación para título
    ws['A1'].font = Font(name="Aptos", size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Fuente y alineación para fecha
    ws['A2'].font = Font(name="Aptos", size=20, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados (fila 3)
    for cell in ws[3]:
        cell.font = Font(name="Aptos", size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de datos (desde la 4)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.font = Font(name="Aptos", size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Wrap text en columnas A, B y C
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colores hex
    FILL_COLORS = {
        "MEAL PACKAGE":   "DCE6F1",
        "BILL TO ROOM":   "EBF1DE",
        "ROOM ONLY":      "F2DCDB",
        "BREAKFAST ONLY": "E4DFEC",
    }

    # Aplicar colores de fondo en la columna "Meal option"
    for i in range(4, ws.max_row + 1):
        meal_value = ws.cell(row=i, column=4).value
        if meal_value in FILL_COLORS:
            ws.cell(row=i, column=4).fill = PatternFill(
                start_color=FILL_COLORS[meal_value],
                end_color=FILL_COLORS[meal_value],
                fill_type="solid"
            )

    thin = Side(style="thin", color="000000")
    all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = all_borders

    wb.save(out_path)

def main_cli():
    parser = argparse.ArgumentParser(description="Generate Meals List from bookings CSV")
    parser.add_argument("--input", required=False, help="Ruta al CSV (bookings report). Si no se pasa, se busca bookings_report*.csv en el cwd.")
    parser.add_argument("--out", required=False, help="Ruta de salida .xlsx")
    args = parser.parse_args()

    # Resolver input
    if args.input:
        csv_path = Path(args.input)
        if not csv_path.exists():
            raise FileNotFoundError(f"No se encontró el CSV en: {csv_path}")
    else:
        # fallback a bookings_report*.csv en el cwd
        candidates = sorted(Path.cwd().glob("bookings_report*.csv"))
        if not candidates:
            raise FileNotFoundError("No se encontró ningún archivo que empiece con 'bookings_report' en la carpeta actual.")
        csv_path = candidates[0]

    # Resolver salida
    date_text = today_str()
    out_path = Path(args.out) if args.out else Path.cwd() / f"meal_list_{date_text}.xlsx"

    print(f"Usando reporte: {csv_path.name}")
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    meals_df = build_meal_list(df)
    write_excel(meals_df, out_path, date_text)
    print(f"✅ Meal list creada: {out_path.name}")

if __name__ == "__main__":
    main_cli()
