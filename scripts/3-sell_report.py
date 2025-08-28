import os
import pandas as pd
from openpyxl import load_workbook
from difflib import get_close_matches

def normalize(text):
    if pd.isna(text):
        return ""
    return (
        str(text)
        .upper()
        .strip()
        .replace("\n", "")
        .replace("\r", "")
        .replace("  ", " ")
    )


# Rutas de archivos
import sys
import shutil

if len(sys.argv) < 2:
    print("❌ You must provide the output path for the final report (e.g. 'report.xlsx').")
    sys.exit(1)

output_path = sys.argv[1]
template_path = "assets/report_template.xlsx"
csv_path = "sale_report.csv"

if not os.path.exists(csv_path):
    print("❌ sale_report.csv not found.")
    sys.exit(1)

# Copiar plantilla a la ruta de salida
shutil.copy(template_path, output_path)
print(f"📄 Copied template to: {output_path}")


# Cargar único archivo CSV
if not os.path.exists(csv_path):
    print("❌ The file sale_report.csv was not found.")
    exit(1)

df = pd.read_csv(csv_path)

if "Product" not in df.columns or "Quantity" not in df.columns:
    print("❌ The file sale_report.csv does not have the required columns ('Product', 'Quantity').")
    exit(1)

df["Product_norm"] = df["Product"].apply(normalize)
lightspeed_df = df[["Product_norm", "Quantity"]]


# Cargar Excel con estilos
wb = load_workbook(output_path)
not_found = []

# Preprocesar cada hoja
products_index = {}  # Dict para guardar {nombre_normalizado: (sheet, row, col)}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Excepción para BEER ON TAP
    if sheet_name == "BEER ON TAP":
        header_row = 16  # fila 16 (0-based)
        data_start = 17  # fila 17 (0-based)
    else:
        header_row = 2   # fila 3 (0-based)
        data_start = 3   # fila 4 (0-based)

    headers = [cell.value for cell in ws[header_row]]

    try:
        product_col = headers.index("Products") + 1
        sold_qty_col = headers.index("Sold qty") + 1
    except ValueError:
        continue

    for row in range(data_start, ws.max_row + 1):
        prod_val = ws.cell(row=row, column=product_col).value
        norm = normalize(prod_val)
        if norm:
            products_index[norm] = (sheet_name, row, sold_qty_col)
        # Vaciar columna A ("Qty") solo a partir de fila 3 (no tocar encabezado)
        if row >= 3:
            ws.cell(row=row, column=1).value = None  # Columna A ("Qty")

        # Reiniciar columna "Sold qty"
        ws.cell(row=row, column=sold_qty_col).value = 0


# Asignar cantidades desde el CSV combinado
for _, row in lightspeed_df.iterrows():
    name = row["Product_norm"]
    qty = row["Quantity"]

    if name in products_index:
        sheet_name, row_idx, col_idx = products_index[name]
        wb[sheet_name].cell(row=row_idx, column=col_idx).value = qty
    else:
        similar = get_close_matches(name, products_index.keys(), n=1, cutoff=0.85)
        if similar:
            print(f"⚠️ No encontrado: '{name}' — ¿Querías decir: '{similar[0]}'?")
        else:
            not_found.append(name)

# Guardar archivo modificado
wb.save(output_path)

# Mostrar resumen
not_found = list(set(not_found))
print(f"\n🟡 Products not found ({len(not_found)}):")
for p in not_found[:100]:
    print(f"• {p}")
