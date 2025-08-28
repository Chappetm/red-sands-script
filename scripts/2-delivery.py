import re
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

def get_next_thursday():
    today = datetime.today()
    days_ahead = 3 - today.weekday()  # 3 = Thursday (0 = Monday)
    if days_ahead <= 0:
        days_ahead += 7
    next_thursday = today + timedelta(days=days_ahead)
    return next_thursday.strftime("%Y-%m-%d")


# Configuraciones
PRODUCTS_FILE = "/Users/matiaschappet/Documents/Parse and upload/assets/products.xlsx"           # Archivo con múltiples hojas (ALM, COKE, CUB, LION)
INVOICES_ROOT = "/Users/matiaschappet/Documents/Parse and upload/Excel_invoices"          # Carpeta con subcarpetas por proveedor
next_thursday = get_next_thursday()

# Proyecto raíz = carpeta padre de /scripts
PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_FILE = PROJECT_ROOT / f"delivery_checklist_{next_thursday}.xlsx"

# --- Clean up: vaciar PDF_invoices (manteniendo la carpeta) ---
from pathlib import Path
import os

try:
    base_dir = Path(__file__).resolve().parent.parent  # carpeta raíz del proyecto
    pdf_dir = base_dir / "PDF_invoices"
    if pdf_dir.exists():
        # 1) borrar archivos
        for p in pdf_dir.rglob("*"):
            if p.is_file():
                try:
                    p.unlink()
                except Exception as e:
                    print(f"⚠️ Could not delete {p}: {e}")
        # 2) quitar subcarpetas vacías
        for p in sorted(pdf_dir.rglob("*"), reverse=True):
            if p.is_dir():
                try:
                    p.rmdir()
                except OSError:
                    pass
    print("CLEANUP_DONE=PDF_invoices")
except Exception as e:
    print(f"⚠️ Cleanup PDF_invoices failed: {e}")


# Carpetas de PDFs
PDF_INVOICES_ROOT = PROJECT_ROOT / "PDF_invoices"

# Cargar todas las hojas del catálogo de productos y combinarlas
xls = pd.ExcelFile(PRODUCTS_FILE)
all_products = []

for sheet in xls.sheet_names:
    df = xls.parse(sheet)
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    df["product_code"] = df["product_code"].astype(str).str.strip()
    df["product_name"] = df["product_name"].astype(str).str.strip()
    # Expandir múltiples códigos por fila
    expanded_rows = []
    for _, row in df.iterrows():
        codes = [code.strip() for code in re.split(r"[\/,;]", row["product_code"]) if code.strip()]
        for code in codes:
            expanded_rows.append({"product_code": code, "product_name": row["product_name"]})

    all_products.append(pd.DataFrame(expanded_rows))

products_df = pd.concat(all_products, ignore_index=True)

# Preparar Excel de salida
wb = Workbook()
wb.remove(wb.active)

# Iterar por cada proveedor
root_dir = Path(INVOICES_ROOT)
for supplier_folder in root_dir.iterdir():
    if supplier_folder.is_dir():
        supplier_name = supplier_folder.name
        supplier_data = []

        # Iterar por archivos de Excel dentro de cada carpeta
        for file in supplier_folder.glob("*.xlsx"):
            if file.name.startswith("~$"):
                continue  # Ignorar archivos temporales de Excel

            print(f"🔍 Trying to read: {file.name}")  # 👈 esto te mostrará qué archivo está por procesar
            df = pd.read_excel(file, engine="openpyxl")
            df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]

            if "product_code" in df.columns and "order_qty" in df.columns:
                subset = df[["product_code", "order_qty"]].copy()
                subset["product_code"] = subset["product_code"].astype(str).str.strip()
                supplier_data.append(subset)

        # Si hay datos, procesar
        if supplier_data:
            combined = pd.concat(supplier_data, ignore_index=True)

            # Combinar con catálogo
            merged = pd.merge(
                combined,
                products_df,
                how="left",
                left_on="product_code",
                right_on="product_code"
            )

            # Detectar códigos no encontrados
            unknown_codes = merged[merged["product_name"].isna()]["product_code"].unique()

            if len(unknown_codes) > 0:
                print(f"🟡 Unknown codes in {supplier_name} ({len(unknown_codes)}):")
                for code in unknown_codes:
                    print(f"• {code}")

            # Ordenar por nombre de producto
            merged["product_name"] = merged["product_name"].fillna("DESCONOCIDO")
            merged = merged.sort_values("product_name")

            # Crear checklist
            checklist_df = pd.DataFrame({
                "Recibido": ["" for _ in range(len(merged))],
                "Producto": merged["product_name"],
                "Cantidad": merged["order_qty"]
            })


            # Agregar hoja al Excel
            ws = wb.create_sheet(title=supplier_name[:31])
            for r_idx, row in enumerate(dataframe_to_rows(checklist_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
            
            # Validación de datos en columna A (Recibido): Sí / No
            dv = DataValidation(type="list", formula1='"✅,❌"', allow_blank=True)
            ws.add_data_validation(dv)
            for row in range(2, len(checklist_df) + 2):
                dv.add(ws[f"A{row}"])

            # Formato condicional: si la celda en A es "Sí", toda la fila se pone verde claro
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            formula = '$A2="✅"'
            rule = FormulaRule(formula=[formula], fill=fill)
            ws.conditional_formatting.add(f"A2:C{len(checklist_df)+1}", rule)

#Si no hay archivos excel de invoices
if not wb.sheetnames:
    print("❌ No valid Excel files found in the supplier folders.")
    exit()

# Guardar archivo final
wb.save(OUTPUT_FILE)
print(f"✅ New file: {OUTPUT_FILE}")
print(f"OUTPUT_FILE={OUTPUT_FILE}")

# 2.2) Eliminar Excel generados por proveedor (Excel_invoices/<supplier>/*.xlsx)
root_dir = Path(INVOICES_ROOT)
deleted_count = 0
if root_dir.exists():
    for supplier_folder in root_dir.iterdir():
        if supplier_folder.is_dir():
            for x in supplier_folder.glob("*.xlsx"):
                if x.name.startswith("~$"):
                    continue  # evitar archivos temporales/abiertos
                try:
                    x.unlink()
                    deleted_count += 1
                    print(f"🧹 Deleted {supplier_folder.name}/{x.name}")
                except Exception as e:
                    print(f"⚠️ Could not delete {supplier_folder.name}/{x.name}: {e}")
