import argparse
import os
import sys
from typing import List, Tuple, Optional
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Categorías que NO se deben modificar desde el script
SKIP_CATEGORIES = {"Beer on tap"}


# ---------------------------
# Config / mappings
# ---------------------------
HEADERS_PRODUCTS = ["code", "name_with_unit"]

CATEGORY_SHEET_MAP = {
    "Beers": "BEER",
    "Beer on tap": "BEER ON TAP",
    "Ciders": "CIDER",
    "RTDs": "RTDS",
    "Wines": "WINE",
    "Spirits": "SPIRITS",
    "Soft drinks": "SOFT",
    "Snacks": "SNACKS",
}

# ---------------------------
# Helpers
# ---------------------------

# ===== Colores por unidad (completar con tus hex) =====
# Acepta "#RRGGBB" o "RRGGBB"; internamente lo convertimos a ARGB "FFRRGGBB".
UNIT_COLOR_HEX = {
    # Cans
    "C1":   "#ECCFED", 
    "C3":   "#BFE5A8",
    "C4":   "#BFE5A8",
    "C6":   "#BFE5A8",
    "C10":  "#C7E5F3",
    "C12":  "#BFE5A8",
    "C16":  "",
    "C20":  "",
    "C24":  "",
    "C30":  "",
    # Stubbies
    "S1":   "#ECCFED",
    "S4":   "#BFE5A8",
    "S6":   "#BFE5A8",
    "S10":  "#C7E5F3",
    "S12":  "#BFE5A8",
    "S24":  "",
    # Bottle / Spirits
    "700ML":"",
    "1L":   "",
}

from openpyxl.styles import PatternFill
import re

def _to_argb(hex_or_hash: str) -> str | None:
    if not hex_or_hash:
        return None
    c = hex_or_hash.strip().lstrip("#")
    if re.fullmatch(r"[0-9A-Fa-f]{6}", c):
        return "FF" + c.upper()         # RRGGBB -> FFRRGGBB
    if re.fullmatch(r"[0-9A-Fa-f]{8}", c):
        return c.upper()                # ya ARGB
    return None

def make_solid_fill(hex_or_hash: str) -> PatternFill | None:
    argb = _to_argb(hex_or_hash)
    if not argb:
        return None
    # Para 'solid' usar fgColor
    return PatternFill(fill_type="solid", fgColor=argb)

def get_unit_fill(unit: str) -> PatternFill | None:
    """Devuelve PatternFill; si el color es vacío/’none’ => None (se limpia)."""
    key = (unit or "").strip().upper()
    raw = (UNIT_COLOR_HEX.get(key, "") or "").strip().lower()
    if raw in ("", "none", "transparent"):
        return None  # => transparente
    # exacto
    fill = make_solid_fill(raw)
    if fill:
        return fill
    # fallback por número (S6 -> C6 si no definido, y viceversa)
    if (key.startswith("C") or key.startswith("S")) and key[1:].isdigit():
        num = key[1:]
        for alt in (f"C{num}", f"S{num}"):
            raw_alt = (UNIT_COLOR_HEX.get(alt, "") or "").strip().lower()
            if raw_alt in ("", "none", "transparent"):
                return None
            fill = make_solid_fill(raw_alt)
            if fill:
                return fill
    return None

def apply_fill_to_row(ws, row: int, fill: PatternFill | None, cols=range(2, 10)):
    """
    Si fill es None -> deja la fila **transparente** (limpia el color).
    Si fill es PatternFill -> aplica el color en B..I.
    """
    if fill is None:
        clear = PatternFill(fill_type=None)
        for col in cols:
            ws.cell(row=row, column=col).fill = clear
        return
    for col in cols:
        ws.cell(row=row, column=col).fill = fill



def find_detail_start(ws, default_start=3):
    for r in range(1, (ws.max_row or 1) + 1):
        val = ws.cell(row=r, column=2).value
        if isinstance(val, str) and val.strip().upper() == "PRODUCTS":
            return r + 1
    return default_start

def first_empty_row(ws, start_row=None, cols=range(2, 10)):
    """Devuelve la primera fila totalmente vacía (B..I) a partir de start_row."""
    if start_row is None:
        start_row = find_detail_start(ws)
    r = start_row
    while True:
        if all(ws.cell(row=r, column=c).value in (None, "") for c in cols):
            return r
        r += 1

def parse_args():
    p = argparse.ArgumentParser(description="Append product into products.xlsx and report.xlsx")
    p.add_argument("-w", "--workbook", required=True, help="Path to products.xlsx")
    p.add_argument("-r", "--report",   required=True, help="Path to report.xlsx")
    p.add_argument("-s", "--supplier", required=True, help="Supplier sheet name in products.xlsx (e.g., ALM)")
    p.add_argument("-c", "--code",     required=True, help="Product code (carton)")
    p.add_argument("-n", "--name",     required=True, help="Product base name (without unit)")
    p.add_argument("--category",       required=True, choices=list(CATEGORY_SHEET_MAP.keys()))
    p.add_argument("--ptype",          required=True, help="Type: 'Cans', 'Stubbies', or 'Bottle'")
    p.add_argument("--units",          required=True, help="Comma-separated units. Eg: C1,C6,C24  |  S1,S24  |  700ML  |  1L")
    p.add_argument("--carton-size", type=int, default=None, help="Carton size (SOFT drinks). Si no se pasa, se asume 24.")

    return p.parse_args()


def open_or_create_xlsx(path: str) -> Workbook:
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    return wb


def ensure_sheet(wb: Workbook, name: str, headers: Optional[List[str]] = None):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(title=name)
        if headers:
            ws.append(headers)
    return ws


def is_header_cell(val: Optional[str], headers: List[str]) -> bool:
    if val is None:
        return False
    if not isinstance(val, str):
        return False
    return val.strip().lower() in {h.lower() for h in headers}


def code_exists_in_sheet(ws, code: str) -> bool:
    code = code.strip()
    for r in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = r[0]
        if val is None:
            continue
        sval = str(val).strip()
        if is_header_cell(sval, HEADERS_PRODUCTS):
            continue
        if sval == code:
            return True
    return False


def save_wb(wb: Workbook, path: str):
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    wb.save(path)


def extract_pack_number(unit: str) -> Optional[int]:
    """'C24' -> 24, 'S6' -> 6, '700ML' -> None"""
    u = unit.strip().upper()
    if len(u) >= 2 and (u[0] in ("C", "S")) and u[1:].isdigit():
        return int(u[1:])
    return None


def is_carton_unit(unit: str) -> bool:
    """Cartón típico si termina en 20/24/30 (podes ampliar si necesitás)"""
    n = extract_pack_number(unit)
    return n in (20, 24, 30)


def alpha_insert_row(ws, product_name: str, start_row: int = 3, col: int = 2) -> int:
    """
    Devuelve la fila en la que hay que insertar product_name para mantener orden alfabético
    comparando contra la columna 'col' (B=2), desde start_row (saltando encabezados).
    """
    target = product_name.strip().upper()
    max_row = ws.max_row or start_row
    for r in range(start_row, max_row + 1):
        cell = ws.cell(row=r, column=col).value
        if not cell:
            # llegó a una fila vacía; insertamos acá
            return r
        cur = str(cell).strip().upper()
        if target < cur:
            return r
    # si no encontró lugar antes, insertamos al final+1
    return max_row + 1


def add_to_products_xlsx(products_path: str, supplier: str, code: str, name: str, unit: str) -> int:
    """
    Agrega [code, name_with_unit] a la hoja del supplier; centra la celda del código.
    Return: 0 OK, 2 duplicado, 1 error.
    """
    try:
        wb = open_or_create_xlsx(products_path)
        ws = ensure_sheet(wb, supplier, headers=HEADERS_PRODUCTS)
    except Exception as e:
        print(f"ERROR: opening products workbook: {e}", file=sys.stderr)
        return 1

    if code_exists_in_sheet(ws, code):
        print(f"WARNING: code already exists in '{supplier}': {code}", file=sys.stderr)
        return 2

    name_with_unit = f"{name.strip()} {unit.strip()}".strip()
    try:
        ws.append([code.strip(), name_with_unit])
        last_row = ws.max_row
        code_cell = ws.cell(row=last_row, column=1)
        code_cell.alignment = Alignment(horizontal="center")
        code_cell.number_format = "@"
        save_wb(wb, products_path)
        print(f"OK products.xlsx: [{code}] {name_with_unit}")
        return 0
    except Exception as e:
        print(f"ERROR: writing products workbook: {e}", file=sys.stderr)
        return 1
    

def upsert_sorted_view(wb, src: str, view: str, only_cartons: bool = False):
    """
    Crea/actualiza una hoja 'view' con una vista ordenada alfabéticamente de src (B..I),
    usando fórmulas dinámicas de Excel (SORT / FILTER / LET). No toca la hoja original.
    """
    # Rango de datos de la hoja fuente (ajustá si tu área va más allá de fila 2000)
    rng = f"{src}!B3:I2000"
    if view in wb.sheetnames:
        ws = wb[view]
        ws.delete_rows(1, ws.max_row or 1)  # limpiamos la vista
    else:
        ws = wb.create_sheet(view)

    # Encabezado opcional
    ws["B1"].value = f"Sorted view of {src}"

    # data = columnas B..I; 
    # INDEX(b,,1)=col B (nombre), ,,4=E (pack size), ,,7=H (base), ,,8=I (carton equiv)
    if only_cartons:
        # Solo filas de cartón (E=H) y con I>0; orden por nombre A→Z
        formula = (
            f'=LET(b,{rng}, '
            f'SORT(FILTER(b, (INDEX(b,,1)<>"")*(INDEX(b,,4)=INDEX(b,,7))*(INDEX(b,,8)>0)), 1, TRUE))'
        )
    else:
        # Todas las filas no vacías, orden por nombre A→Z
        formula = f'=LET(b,{rng}, SORT(FILTER(b, INDEX(b,,1)<>""), 1, TRUE))'

    ws["B2"].value = formula



# ---------------------------
# Report logic (inserción con fórmulas)
# ---------------------------
def row_exists(ws, product_name: str, colB: int = 2, start_row: int = 3) -> bool:
    for r in range(start_row, (ws.max_row or start_row) + 1):
        val = ws.cell(row=r, column=colB).value
        if val and str(val).strip().upper() == product_name.strip().upper():
            return True
    return False


def insert_rows(ws, at_row: int, count: int):
    """Inserta 'count' filas a partir de 'at_row' (desplaza hacia abajo)."""
    ws.insert_rows(at_row, amount=count)

def set_soft_row(ws, row: int, name_with_unit: str, carton_size: int):
    """
    SOFT: B=name, C=0, D='unit', E=1, F='=E*C', G='=F/H', H=carton_size
    Igual al template: fila simple, 'G' es el equivalente en cartones.
    """
    ws.cell(row=row, column=2, value=name_with_unit)        # B
    ws.cell(row=row, column=3, value=0)                     # C (Sold qty)
    ws.cell(row=row, column=4, value='unit')                # D (Unit)
    ws.cell(row=row, column=5, value=1)                     # E (DefaultSize)
    ws.cell(row=row, column=6, value=f"=E{row}*C{row}")     # F (Total)
    g = ws.cell(row=row, column=7, value=f"=F{row}/H{row}") # G (Total/carton)
    g.number_format = "0.00"
    ws.cell(row=row, column=8, value=carton_size)           # H (carton size)

def set_snack_row(ws, row: int, name_with_unit: str):
    """
    SNACKS: B=name, C=0, D='unit', E=1, H='=C' (Total)
    """
    ws.cell(row=row, column=2, value=name_with_unit)  # B
    ws.cell(row=row, column=3, value=0)               # C
    ws.cell(row=row, column=4, value='unit')          # D
    ws.cell(row=row, column=5, value=1)               # E (para mantener patrón)
    ws.cell(row=row, column=8, value=f"=C{row}")      # H (Total)


def set_beer_or_cider_row(ws, row: int, name_with_unit: str, pack_size: int, base_carton: int, is_carton: bool):
    """
    Escribe una fila de BEER/CIDER:
      B=name, C=0, D='unit', E=pack, G=... , H=base, I=fórmula por fila (packs)
    """
    ws.cell(row=row, column=2, value=name_with_unit)
    ws.cell(row=row, column=3, value=0)                # Sold qty (input)
    ws.cell(row=row, column=4, value='unit')           # Unit label
    ws.cell(row=row, column=5, value=pack_size)        # Pack size (E)
    # F (col 6) se deja en blanco
    if is_carton:
        ws.cell(row=row, column=7, value=f"=C{row}")   # G = C (cartones)
    else:
        ws.cell(row=row, column=7, value=f"=C{row}*E{row}")  # G = C*E
    ws.cell(row=row, column=8, value=base_carton)      # H = base de familia

    if not is_carton:
        cell_I = ws.cell(row=row, column=9, value=f"=G{row}/H{row}")  # I packs
        cell_I.number_format = "0.00"
    # Para cartón, I se setea después con set_cider_or_beer_carton_I


def set_cider_or_beer_carton_I(ws, carton_row: int, pack_rows: List[int]):
    """
    En BEER/CIDER: I(cartón) = G(cartón) + sum(I de packs)
    """
    parts = [f"G{carton_row}"] + [f"I{r}" for r in pack_rows]
    formula = "=" + "+".join(parts)
    cell_I = ws.cell(row=carton_row, column=9, value=formula)

    # Formato numérico y estilo "rojo final"
    cell_I.number_format = "0.00"
    cell_I.font = Font(color="FF0000", bold=True, size=16)



def set_spirits_row(ws, row: int, name_with_unit: str, ml: int, type_label: Optional[str] = None):
    """
    SPIRITS: B=name, C=0, D='ml', E=ml, G='=Crow', H='=Grow', I=type (opcional)
    """
    ws.cell(row=row, column=2, value=name_with_unit)
    ws.cell(row=row, column=3, value=0)
    ws.cell(row=row, column=4, value='ml')
    ws.cell(row=row, column=5, value=ml)
    ws.cell(row=row, column=7, value=f"=C{row}")    # G
    ws.cell(row=row, column=8, value=f"=G{row}")    # H
    if type_label:
        ws.cell(row=row, column=9, value=type_label)


def set_wine_row(ws, row: int, name_with_unit: str):
    """
    WINE: B=name, C=0, D='ml', E=750, G='=Crow'
    (En tu template hay 10 columnas; dejamos Type en blanco)
    """
    ws.cell(row=row, column=2, value=name_with_unit)
    ws.cell(row=row, column=3, value=0)
    ws.cell(row=row, column=4, value='ml')
    ws.cell(row=row, column=5, value=750)
    ws.cell(row=row, column=7, value=f"=C{row}")    # G
    # H/I/J se dejan como en el template (vacíos por defecto)


def update_report(report_path: str, name: str, category: str, ptype: str, units_csv: str) -> Tuple[int, str]:
    """
    Inserta filas en la planilla de report según categoría y unidades.
    Devuelve (status_code, message). status_code: 0 OK, 3 todo duplicado, 1 error.
    """

    if category in SKIP_CATEGORIES:
        return 0, f"Skipped category '{category}' (no changes by policy)"
    
    wb = open_or_create_xlsx(report_path)

    if category not in CATEGORY_SHEET_MAP:
        return 1, f"Unknown category '{category}'"
    sheet_name = CATEGORY_SHEET_MAP[category]
    if sheet_name not in wb.sheetnames:
        return 1, f"Sheet '{sheet_name}' not found in report"

    ws = wb[sheet_name]

    # Normalizamos unidades
    raw_units = [u.strip().upper() for u in units_csv.split(",") if u.strip()]
    # Para Wine, usar S1 (internamente 750ml); para Spirits usar 700ML/1L.
    # Evitar filas terminadas en 'CAN', 'SUBBIE', 'NIP' (regla dada).
    blocked_suffixes = (" CAN", " SUBBIE", " NIP")

    created_rows = []

    if category in ("Beers", "Ciders"):
        # Prefijo según tipo
        if ptype not in ("Cans", "Stubbies"):
            return 1, f"For {category}, ptype must be Cans or Stubbies"
        pref = "C" if ptype == "Cans" else "S"

        # Filtrar a solo unidades C#/S#
        units = []
        for u in raw_units:
            if u.startswith(pref) and extract_pack_number(u) is not None:
                units.append(u)

        if not units:
            return 1, "No valid units for BEER/CIDER"

        # --- Unidades válidas (C#/S#) ---
        units = []
        for u in raw_units:
            if u.startswith(pref) and extract_pack_number(u) is not None:
                units.append(u)
        if not units:
            return 1, "No valid units for BEER/CIDER"

        # --- Cartón = máximo pack (soporta 16/20/24/30/...) ---
        nums = [extract_pack_number(u) for u in units]
        nums = [n for n in nums if n is not None and n > 0]
        if not nums:
            return 1, "No valid numeric units for BEER/CIDER"
        base = max(nums)
        carton_unit = f"{pref}{base}"

        # Orden lógico: packs chicos→grandes y el cartón al final (aunque no venga en units, lo agregamos)
        units_sorted = sorted(units, key=lambda x: extract_pack_number(x))
        if carton_unit in units_sorted:
            units_sorted = [u for u in units_sorted if u != carton_unit] + [carton_unit]
        else:
            units_sorted.append(carton_unit)

        # ------- APPEND ONLY: escribir desde la primera fila vacía -------
        start = find_detail_start(ws)
        insert_at = first_empty_row(ws, start_row=start)

        # Escribir filas + fórmulas
        pack_rows = []
        carton_row = None
        for idx, u in enumerate(units_sorted):
            r = insert_at + idx
            n = extract_pack_number(u) or 0
            is_carton = (u == carton_unit)

            set_beer_or_cider_row(
                ws, r,
                name_with_unit=f"{name} {u}",
                pack_size=n,
                base_carton=base,
                is_carton=is_carton,
            )

            # Color por unidad
            apply_fill_to_row(ws, r, get_unit_fill(u))

            if is_carton:
                carton_row = r
            else:
                pack_rows.append(r)
            created_rows.append((ws.title, r, f"{name} {u}"))

        # Fórmula final en I(cartón) + formato
        if carton_row is not None:
            set_cider_or_beer_carton_I(ws, carton_row, pack_rows)


    elif category == "Wines":
        unit = "S1"
        row_name = f"{name} {unit}"
        start = find_detail_start(ws)
        insert_at = first_empty_row(ws, start_row=start)

        if not row_exists(ws, row_name, start_row=start):
            set_wine_row(ws, insert_at, row_name)
            apply_fill_to_row(ws, insert_at, get_unit_fill("S1"))
            created_rows.append((ws.title, insert_at, row_name))
        else:
            print(f"Report: row already exists '{row_name}'", file=sys.stderr)

    elif category == "Spirits":
        normalized = raw_units[0] if raw_units else ""
        if normalized not in ("700ML", "1L"):
            return 1, "Spirits must be 700ML or 1L"
        ml = 700 if normalized == "700ML" else 1000
        row_name = f"{name} {normalized}"

        start = find_detail_start(ws)
        insert_at = first_empty_row(ws, start_row=start)

        if not row_exists(ws, row_name, start_row=start):
            set_spirits_row(ws, insert_at, row_name, ml=ml, type_label=None)
            suffix = "700ML" if ml == 700 else "1L"
            apply_fill_to_row(ws, insert_at, get_unit_fill(suffix))
            created_rows.append((ws.title, insert_at, row_name))
        else:
            print(f"Report: row already exists '{row_name}'", file=sys.stderr)

    elif category == "RTDs":
        if ptype not in ("Cans", "Stubbies"):
            return 1, "For RTDs, ptype must be Cans or Stubbies"
        pref = "C" if ptype == "Cans" else "S"

        units = []
        for u in raw_units:
            if u.startswith(pref) and extract_pack_number(u) is not None:
                units.append(u)
        if not units:
            return 1, "No valid units for RTDs"

        nums = [extract_pack_number(u) for u in units]
        nums = [n for n in nums if n]
        base = max(nums)
        carton_unit = f"{pref}{base}"

        units_sorted = sorted(units, key=lambda x: extract_pack_number(x))
        if carton_unit in units_sorted:
            units_sorted = [u for u in units_sorted if u != carton_unit] + [carton_unit]
        else:
            units_sorted.append(carton_unit)

        start = find_detail_start(ws)
        insert_at = first_empty_row(ws, start_row=start)

        pack_rows, carton_row = [], None
        for i, u in enumerate(units_sorted):
            r = insert_at + i
            n = extract_pack_number(u) or 0
            is_carton = (u == carton_unit)

            set_beer_or_cider_row(ws, r, f"{name} {u}", n, base, is_carton)
            apply_fill_to_row(ws, r, get_unit_fill(u))

            if not is_carton:
                cell = ws.cell(row=r, column=9)  # I packs
                cell.number_format = "0.00"
                pack_rows.append(r)
            else:
                carton_row = r

        if carton_row:
            set_cider_or_beer_carton_I(ws, carton_row, pack_rows)
            ci = ws.cell(row=carton_row, column=9)
            ci.number_format = "0.00"
            ci.font = Font(color="FF0000", bold=True, size=16)

    elif category == "Soft drinks":
        # carton_size: si te pasan --carton-size lo usás, si no default 24
        carton_size = 24
        try:
            # Si tu main() le pasa como parámetro a update_report, usalo;
            # si no, toma override por env (mantiene compat):
            carton_size = int(os.getenv("SOFT_CARTON_SIZE_OVERRIDE", carton_size))
        except Exception:
            pass

        unit = raw_units[0] if raw_units else "S1"
        row_name = f"{name} {unit}".strip()

        start = find_detail_start(ws)
        at = first_empty_row(ws, start_row=start)

        if not row_exists(ws, row_name, start_row=start):
            set_soft_row(ws, at, row_name, carton_size=carton_size)
            apply_fill_to_row(ws, at, get_unit_fill("S1"))  # si querés color base
            created_rows.append((ws.title, at, row_name))
        else:
            print(f"Report: row already exists '{row_name}'", file=sys.stderr)

    elif category == "Snacks":
        row_name = name.strip()

        start = find_detail_start(ws)
        at = first_empty_row(ws, start_row=start)

        if not row_exists(ws, row_name, start_row=start):
            set_snack_row(ws, at, row_name)
            created_rows.append((ws.title, at, row_name))
        else:
            print(f"Report: row already exists '{row_name}'", file=sys.stderr)
    

def main():
    args = parse_args()

    # 1) products.xlsx
    ret_products = add_to_products_xlsx(
        products_path=args.workbook,
        supplier=args.supplier.strip(),
        code=args.code.strip(),
        name=args.name.strip(),
        unit=args.units.split(",")[0].strip()  # para products usamos la unidad "principal"
    )
    if ret_products not in (0, 2):
        sys.exit(1)  # error duro

    # 2) report.xlsx
    ret_report, msg = update_report(
        report_path=args.report,
        name=args.name.strip(),
        category=args.category.strip(),
        ptype=args.ptype.strip(),
        units_csv=args.units.strip()
    )
    if ret_report == 0:
        print(msg)
        sys.exit(0 if ret_products == 0 else 2)  # 0 OK ambos; 2 si dup en products pero report OK
    elif ret_report == 3:
        print(msg, file=sys.stderr)
        sys.exit(0 if ret_products == 0 else 2)
    else:
        print(msg, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    sys.exit(main())
